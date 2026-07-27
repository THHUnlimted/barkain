"""Distributor price-list ingest — CSV / TSV / XLSX → normalized row drafts.

Pure module: bytes in, dataclasses out. No DB, no network.

## The problem

There is no standard for distributor price lists. Across a dozen suppliers
you'll see the UPC column called ``UPC``, ``UPC Code``, ``GTIN``, ``Barcode``,
``EAN``, ``Item UPC`` or ``U.P.C.``; the cost column called ``Cost``, ``Unit
Cost``, ``Your Price``, ``Dealer``, ``Wholesale``, ``WSP`` or ``Net``; and the
whole thing prefixed by four rows of company letterhead before the real header.

So ingest does three things nobody enjoys writing:

1. **Find the header row.** Not always row 1. We score the first
   ``_HEADER_SCAN_ROWS`` rows and take the one that looks most like a header.
2. **Map columns by meaning, not position.** Keyword scoring against a synonym
   table, best-match-wins, each target column claimed at most once.
3. **Parse money and integers out of human text.** ``$1,234.56``, ``(2.50)``
   for negatives, ``12.5 USD``, ``1,200``, ``12 ea``, ``—`` for null.

## Case-pack economics

The single most common way to get a wholesale spreadsheet wrong is to compare a
*case* cost against an *each* selling price and conclude everything is a loser.
When a list gives a case cost and a case pack, the unit cost is
``case_cost / case_pack`` — and if it gives both a unit cost and a case cost we
prefer the explicit unit cost and record the disagreement rather than silently
picking one.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

from m15_sourcing.upc import NormalizedUpc, normalize_upc

# MARK: - Column synonym table
#
# Order within each tuple doesn't matter; scoring is by best match. Entries are
# lowercased substrings matched against a normalized header cell. Keep the
# *specific* variants ahead of generic ones in the code that consumes this —
# "case cost" must not be claimed by the "cost" rule.

COLUMN_SYNONYMS: dict[str, tuple[str, ...]] = {
    "upc": (
        "upc", "u.p.c", "gtin", "barcode", "bar code", "ean", "item upc",
        "upc code", "upca", "product code", "scan code",
    ),
    "case_cost": (
        "case cost", "case price", "cost per case", "price per case",
        "master case cost", "case", "cs cost", "case $",
    ),
    "unit_cost": (
        "unit cost", "cost per unit", "your price", "your cost", "dealer price",
        "dealer cost", "wholesale price", "wholesale cost", "wholesale", "wsp",
        "net price", "net cost", "each cost", "cost each", "ea cost",
        "unit price", "invoice cost", "buy price", "cost",
    ),
    "case_pack": (
        "case pack", "casepack", "pack size", "units per case", "qty per case",
        "qty/case", "units/case", "pack qty", "inner pack", "pack", "upc/case",
    ),
    "moq": (
        "moq", "min order", "minimum order", "min qty", "minimum qty",
        "min order qty", "order minimum",
    ),
    "msrp": (
        "msrp", "map", "retail price", "srp", "list price", "suggested retail",
        "retail",
    ),
    "description": (
        "description", "item description", "product description", "product name",
        "item name", "title", "product", "item",
    ),
    "brand": ("brand", "manufacturer", "mfr", "mfg", "vendor", "supplier", "make"),
    "sku": ("sku", "item number", "item #", "item no", "part number", "part #",
            "model", "mpn", "manufacturer part", "vendor sku"),
    "lead_time_days": (
        "lead time", "leadtime", "lead days", "ship time", "ships in",
        "restock days", "replenishment", "eta days",
    ),
    "quantity_available": (
        "qty available", "available", "on hand", "stock", "inventory",
        "qty on hand", "avail",
    ),
}

# Column order to resolve in. More specific targets claim their header first so
# a "Case Cost" header can't be swallowed by the generic "cost" synonym under
# `unit_cost`. This ordering is load-bearing — see `_map_columns`.
_RESOLUTION_ORDER: tuple[str, ...] = (
    "upc", "case_pack", "case_cost", "unit_cost", "moq", "msrp",
    "lead_time_days", "brand", "sku", "quantity_available", "description",
)

_HEADER_SCAN_ROWS = 15
_MAX_PREVIEW_ROWS = 5

_MONEY_CLEAN_RE = re.compile(r"[^\d.\-()]")
_INT_CLEAN_RE = re.compile(r"[^\d\-]")
_NULLISH = {"", "-", "--", "—", "n/a", "na", "null", "none", "tbd", "."}


# MARK: - Value parsing


def parse_money(raw: object) -> Decimal | None:
    """Parse a currency-ish cell into a ``Decimal``, or ``None``.

    Handles ``$1,234.56``, ``1234.56 USD``, ``(2.50)`` (accounting negative),
    bare floats, and the usual null placeholders. Returns ``None`` rather than
    raising — a distributor list with one junk cell shouldn't fail the upload.
    """
    if raw is None:
        return None
    if isinstance(raw, (int, float, Decimal)):
        try:
            return Decimal(str(raw))
        except InvalidOperation:
            return None

    text = str(raw).strip().lower()
    if text in _NULLISH:
        return None

    negative = text.startswith("(") and text.endswith(")")
    cleaned = _MONEY_CLEAN_RE.sub("", text).replace("(", "").replace(")", "")
    if not cleaned or cleaned in {"-", ".", "-."}:
        return None
    try:
        value = Decimal(cleaned)
    except InvalidOperation:
        return None
    return -value if negative and value > 0 else value


def parse_int(raw: object) -> int | None:
    """Parse an integer out of a cell like ``"12 ea"``, ``"1,200"`` or ``12.0``."""
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, int):
        return raw
    if isinstance(raw, (float, Decimal)):
        return int(raw)

    text = str(raw).strip().lower()
    if text in _NULLISH:
        return None
    # A cell like "12.0" should read as 12, not 120 once we strip the dot.
    if "." in text:
        head, _, tail = text.partition(".")
        if tail.strip(" 0") == "":
            text = head
    cleaned = _INT_CLEAN_RE.sub("", text)
    if not cleaned or cleaned == "-":
        return None
    try:
        return int(cleaned)
    except ValueError:
        return None


def _clean_text(raw: object) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text or text.lower() in _NULLISH:
        return None
    return re.sub(r"\s+", " ", text)


# MARK: - Header detection


def _normalize_header(cell: object) -> str:
    return re.sub(r"[^a-z0-9 /#$.]+", " ", str(cell or "").strip().lower()).strip()


def _score_header_cell(header: str, synonyms: tuple[str, ...]) -> int:
    """Score how well one header cell matches a synonym set.

    Exact match beats prefix beats substring. Longer synonyms score higher so
    ``"case cost"`` outranks ``"cost"`` on the cell ``"Case Cost"`` — that
    length weighting is what keeps case and unit columns from swapping.
    """
    if not header:
        return 0
    best = 0
    for syn in synonyms:
        if header == syn:
            score = 1000 + len(syn)
        elif header.startswith(syn) or header.endswith(syn):
            score = 500 + len(syn)
        elif syn in header:
            score = 200 + len(syn)
        else:
            continue
        best = max(best, score)
    return best


def _looks_like_header(row: list[str]) -> int:
    """Heuristic score for "is this the header row".

    A header row has several non-empty, mostly-non-numeric cells that match our
    synonym table. Letterhead rows ("ACME Distributing, effective 1/1/26")
    have one populated cell; data rows are numeric-heavy.
    """
    cells = [_normalize_header(c) for c in row]
    populated = [c for c in cells if c]
    if len(populated) < 2:
        return 0

    numeric = sum(1 for c in populated if re.fullmatch(r"[\d.,$%-]+", c))
    if numeric > len(populated) / 2:
        return 0

    score = 0
    for target in _RESOLUTION_ORDER:
        syns = COLUMN_SYNONYMS[target]
        if any(_score_header_cell(c, syns) >= 500 for c in cells):
            score += 10
    return score + len(populated)


# MARK: - Results


@dataclass(frozen=True)
class ColumnMapping:
    """Which spreadsheet column index feeds which logical field."""

    header_row_index: int
    headers: tuple[str, ...]
    columns: dict[str, int]  # logical name → column index
    unmapped_headers: tuple[str, ...]

    def get(self, name: str, row: list[object]) -> object | None:
        idx = self.columns.get(name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    @property
    def has_required(self) -> bool:
        """A list without a UPC column and *some* cost column can't be crunched."""
        return "upc" in self.columns and (
            "unit_cost" in self.columns or "case_cost" in self.columns
        )


@dataclass
class IngestedRow:
    """One distributor line, normalized but not yet matched or priced."""

    row_index: int
    raw: dict[str, object]
    upc: NormalizedUpc
    description: str | None = None
    brand: str | None = None
    sku: str | None = None
    unit_cost: Decimal | None = None
    case_cost: Decimal | None = None
    case_pack: int | None = None
    moq: int | None = None
    msrp: Decimal | None = None
    quantity_available: int | None = None
    # Per-product reorder lead time. Supplied by the buyer, either as a column
    # on the price list or set per row afterwards — it varies by supplier far
    # too much for a single global default to be honest.
    lead_time_days: int | None = None
    notes: list[str] = field(default_factory=list)

    @property
    def effective_unit_cost(self) -> Decimal | None:
        """Cost of one sellable unit — the number every downstream calculation wants.

        Prefers an explicit unit cost. Falls back to ``case_cost / case_pack``,
        which is the single most common source of "everything looks unprofitable"
        in a hand-built sourcing spreadsheet.
        """
        if self.unit_cost is not None:
            return self.unit_cost
        if self.case_cost is not None and self.case_pack:
            return self.case_cost / Decimal(self.case_pack)
        return None

    @property
    def minimum_buy_units(self) -> int | None:
        """Units you must commit to on a first order — MOQ, else one case."""
        if self.moq:
            return self.moq
        if self.case_pack:
            return self.case_pack
        return None

    @property
    def minimum_buy_cost(self) -> Decimal | None:
        """Cash outlay for the minimum commitment. Drives the capital-at-risk column."""
        unit = self.effective_unit_cost
        units = self.minimum_buy_units
        if unit is None or units is None:
            return None
        return unit * Decimal(units)


@dataclass(frozen=True)
class IngestResult:
    rows: list[IngestedRow]
    mapping: ColumnMapping
    total_data_rows: int
    skipped_rows: int
    warnings: list[str]

    @property
    def preview(self) -> list[IngestedRow]:
        return self.rows[:_MAX_PREVIEW_ROWS]


class IngestError(ValueError):
    """Raised when a file can't be parsed at all (bad format, no header, no UPC column)."""


# MARK: - Column mapping


def _map_columns(header_row: list[object], header_row_index: int) -> ColumnMapping:
    """Assign each logical field to at most one column, most-specific first.

    Resolution runs in ``_RESOLUTION_ORDER`` and marks a column claimed once
    assigned, so ``Case Cost`` is taken by ``case_cost`` before ``unit_cost``
    ever sees it. Without that ordering, a list with both columns silently maps
    the case price as the unit price — a 12× cost error that reads as plausible.
    """
    normalized = [_normalize_header(c) for c in header_row]
    claimed: set[int] = set()
    columns: dict[str, int] = {}

    for target in _RESOLUTION_ORDER:
        syns = COLUMN_SYNONYMS[target]
        best_idx, best_score = None, 0
        for idx, header in enumerate(normalized):
            if idx in claimed or not header:
                continue
            score = _score_header_cell(header, syns)
            if score > best_score:
                best_idx, best_score = idx, score
        # 200 is the "substring anywhere" floor. Below that we'd be matching on
        # coincidence (a "Discount" column contains "count", not "cost").
        if best_idx is not None and best_score >= 200:
            columns[target] = best_idx
            claimed.add(best_idx)

    unmapped = tuple(
        str(header_row[i] or "") for i in range(len(normalized)) if i not in claimed
    )
    return ColumnMapping(
        header_row_index=header_row_index,
        headers=tuple(str(h or "") for h in header_row),
        columns=columns,
        unmapped_headers=unmapped,
    )


# MARK: - Raw table readers


def _read_delimited(data: bytes) -> list[list[str]]:
    """Decode + sniff the delimiter of a CSV/TSV payload."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:  # pragma: no cover — latin-1 decodes any byte string
        raise IngestError("Could not decode file as text.")

    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        # Sniffer fails on single-column files and on ragged letterhead rows.
        # Fall back to whichever candidate appears most in the sample.
        delimiter = max(",;\t|", key=sample.count)

    return [row for row in csv.reader(io.StringIO(text), delimiter=delimiter)]


def _read_xlsx(data: bytes) -> list[list[object]]:
    """Read the first worksheet of an XLSX via openpyxl (optional dependency)."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover — depends on install profile
        raise IngestError(
            "XLSX support requires the 'openpyxl' package. "
            "Export the list as CSV, or install openpyxl."
        ) from exc

    try:
        # read_only + data_only: we want computed values, not formulas, and we
        # never want a 3,000-row workbook fully materialized as cell objects.
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise IngestError(f"Could not open the workbook: {exc}") from exc

    try:
        sheet = workbook[workbook.sheetnames[0]]
        return [list(row) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def read_table(data: bytes, filename: str = "") -> list[list[object]]:
    """Dispatch on file extension, sniffing content when the name doesn't say."""
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xlsm", ".xltx")):
        return _read_xlsx(data)
    if lower.endswith((".csv", ".tsv", ".txt")):
        return [list(r) for r in _read_delimited(data)]
    # Unknown extension: XLSX is a zip, so the PK magic number is decisive.
    if data[:2] == b"PK":
        return _read_xlsx(data)
    return [list(r) for r in _read_delimited(data)]


# MARK: - Entry point


def ingest(
    data: bytes,
    filename: str = "",
    *,
    max_rows: int = 10_000,
    default_case_pack: int | None = None,
    cost_is_case_cost: bool = False,
) -> IngestResult:
    """Parse a distributor price list into normalized rows.

    ``cost_is_case_cost`` is the manual override for lists whose single cost
    column is per-case but labelled ambiguously ("Price"). ``default_case_pack``
    fills in a pack size for lists that state it in prose ("all items 12/case")
    rather than in a column.

    Raises ``IngestError`` when the file has no usable header or no UPC column —
    those are upload-time failures, not per-row ones.
    """
    table = read_table(data, filename)
    if not table:
        raise IngestError("The file is empty.")

    warnings: list[str] = []

    # Find the header row.
    best_idx, best_score = 0, -1
    for idx, row in enumerate(table[:_HEADER_SCAN_ROWS]):
        score = _looks_like_header([str(c or "") for c in row])
        if score > best_score:
            best_idx, best_score = idx, score
    if best_score <= 0:
        raise IngestError(
            "Couldn't find a header row. Expected a row with column names like "
            "'UPC' and 'Cost' in the first 15 rows."
        )
    if best_idx > 0:
        warnings.append(
            f"Skipped {best_idx} row(s) above the header (letterhead or notes)."
        )

    mapping = _map_columns(table[best_idx], best_idx)
    if "upc" not in mapping.columns:
        raise IngestError(
            "No UPC column found. Looked for: UPC, GTIN, Barcode, EAN, Item UPC."
        )
    if not mapping.has_required:
        raise IngestError(
            "No cost column found. Looked for: Cost, Unit Cost, Your Price, "
            "Wholesale, Case Cost."
        )

    if cost_is_case_cost and "unit_cost" in mapping.columns and "case_cost" not in mapping.columns:
        # Caller says the mapped "cost" is really per-case; move the claim.
        mapping.columns["case_cost"] = mapping.columns.pop("unit_cost")

    rows: list[IngestedRow] = []
    skipped = 0
    truncated = False

    for offset, raw_row in enumerate(table[best_idx + 1 :], start=1):
        if len(rows) >= max_rows:
            truncated = True
            break
        if not any(str(c or "").strip() for c in raw_row):
            continue  # blank spacer row

        upc_cell = mapping.get("upc", raw_row)
        normalized = normalize_upc(upc_cell)

        row = IngestedRow(
            row_index=best_idx + offset,
            raw={
                str(mapping.headers[i]) if i < len(mapping.headers) else f"col_{i}": (
                    raw_row[i] if not isinstance(raw_row[i], (bytes, bytearray)) else None
                )
                for i in range(len(raw_row))
            },
            upc=normalized,
            description=_clean_text(mapping.get("description", raw_row)),
            brand=_clean_text(mapping.get("brand", raw_row)),
            sku=_clean_text(mapping.get("sku", raw_row)),
            unit_cost=parse_money(mapping.get("unit_cost", raw_row)),
            case_cost=parse_money(mapping.get("case_cost", raw_row)),
            case_pack=parse_int(mapping.get("case_pack", raw_row)) or default_case_pack,
            moq=parse_int(mapping.get("moq", raw_row)),
            msrp=parse_money(mapping.get("msrp", raw_row)),
            quantity_available=parse_int(mapping.get("quantity_available", raw_row)),
            lead_time_days=parse_int(mapping.get("lead_time_days", raw_row)),
        )

        # A row with neither a usable UPC nor a description is noise — a footer,
        # a subtotal line, a page break. Dropping it keeps the row count honest.
        if not row.upc.is_usable and not row.description:
            skipped += 1
            continue

        if row.unit_cost is not None and row.case_cost is not None and row.case_pack:
            derived = row.case_cost / Decimal(row.case_pack)
            # 1% tolerance absorbs rounding in the distributor's own math.
            if derived and abs(derived - row.unit_cost) / derived > Decimal("0.01"):
                row.notes.append(
                    f"unit_cost ({row.unit_cost}) disagrees with case_cost/case_pack "
                    f"({derived:.4f}); using unit_cost"
                )

        if row.effective_unit_cost is None:
            row.notes.append("no usable cost — row cannot be scored")

        rows.append(row)

    if truncated:
        warnings.append(
            f"List truncated at {max_rows} rows. Split the file to crunch the rest."
        )

    unusable = sum(1 for r in rows if not r.upc.is_usable)
    if rows and unusable / len(rows) > 0.5:
        warnings.append(
            f"{unusable} of {len(rows)} rows have an unusable UPC — check that the "
            f"UPC column mapped correctly (matched '{mapping.headers[mapping.columns['upc']]}')."
        )

    return IngestResult(
        rows=rows,
        mapping=mapping,
        total_data_rows=len(table) - best_idx - 1,
        skipped_rows=skipped,
        warnings=warnings,
    )
